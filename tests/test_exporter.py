import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from metacritic_scraper_py.exporter import export_sqlite_to_excel
from metacritic_scraper_py.storage import SQLiteStorage


class ExporterTestCase(unittest.TestCase):
    def test_export_sqlite_to_excel(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            db_path = base / "test.db"
            xlsx_path = base / "out.xlsx"

            storage = SQLiteStorage(db_path)
            try:
                storage.upsert_game(
                    slug="demo-game",
                    product_payload={
                        "data": {
                            "item": {
                                "id": 1,
                                "title": "Demo Game",
                                "platform": "PC",
                                "releaseDate": "2026-03-05",
                                "premiereYear": 2026,
                                "rating": "T",
                            }
                        }
                    },
                    critic_summary_payload={"data": {"item": {"score": 85, "reviewCount": 10}}},
                    user_summary_payload={"data": {"item": {"score": 8.2, "reviewCount": 12}}},
                )
                storage.upsert_critic_reviews(
                    "demo-game",
                    [
                        {
                            "publicationSlug": "demo-publication",
                            "publicationName": "Demo Publication",
                            "date": "2026-03-05",
                            "score": 80,
                            "url": "https://example.com/review",
                            "quote": "Good game.",
                            "author": "Critic",
                        }
                    ],
                )
                storage.upsert_user_reviews(
                    "demo-game",
                    [
                        {
                            "id": "user-review-1",
                            "author": "UserA",
                            "score": 9,
                            "date": "2026-03-05",
                            "spoiler": False,
                            "quote": "Great game.",
                        }
                    ],
                )
            finally:
                storage.close()

            counts = export_sqlite_to_excel(db_path=db_path, output_path=xlsx_path)
            self.assertEqual(counts["games_rows"], 1)
            self.assertEqual(counts["critic_reviews_rows"], 1)
            self.assertEqual(counts["user_reviews_rows"], 1)
            self.assertTrue(xlsx_path.exists())

            workbook = load_workbook(xlsx_path, read_only=True)
            try:
                self.assertIn("Summary", workbook.sheetnames)
                self.assertIn("Games", workbook.sheetnames)
                self.assertIn("CriticReviews", workbook.sheetnames)
                self.assertIn("UserReviews", workbook.sheetnames)
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()

